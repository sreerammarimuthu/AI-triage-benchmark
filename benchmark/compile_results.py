"""
compile_results.py — Generates results for the primary prompt format.

Sheets:
  1.  Summary           All 5 models side-by-side (key metrics)
  2.  Predictions       Every model × case prediction (390 rows: 5 models × 78 cases)
  3.  By_Model          Aggregate metrics per model
  4.  By_PromptType     E vs F accuracy per model
  5.  By_Domain         Per-domain accuracy per model
  6.  By_Triage         Per gold-triage-level metrics per model
  7.  Hallucination     Coherence + faithfulness rates per model
  8.  Halluc_Breakdown  Per-case hallucination details (all models)
  9.  Under_Triage      Safety sheet: under-triage cases (pred < min gold)
  10. Over_Triage       Efficiency sheet: over-triage cases (pred > max gold)
  11. Clear_vs_Edge     All metrics per model for clear vs edge cases side by side
  12. EF_by_Triage      E vs F accuracy + faithfulness per gold triage level per model
  13. EF_Faithfulness   E vs F faithfulness/coherence delta — hallucination-amplification finding
  14. Confusion_[model] One confusion matrix per model (5 sheets)

"""

import json
import math
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────
CKPT_V2   = "benchmark/output/results/checkpoint.json"
OUT_PATH  = "benchmark/output/reports/results.xlsx"

# ── Constants ───────────────────────────────────────────────────────────────
TRIAGE_ORD = {"A": 1, "B": 2, "C": 3, "D": 4}

MODEL_IDS = [
    "mistral:latest",
    "deepseek-r1:7b",
    "gemma2:9b",
    "qwen2.5:7b",
    "chatgpt-health",
]
MODEL_DISPLAY = {
    "mistral:latest":  "Mistral 7B",
    "deepseek-r1:7b":  "DeepSeek-R1 7B",
    "gemma2:9b":       "Gemma2 9B",
    "qwen2.5:7b":      "Qwen2.5 7B",
    "chatgpt-health":  "ChatGPT Health",
}
MODEL_COLORS = {
    "mistral:latest":  "E3F2FD",  # pale blue
    "deepseek-r1:7b":  "FCE4EC",  # pale pink
    "gemma2:9b":       "E8F5E9",  # pale green
    "qwen2.5:7b":      "FFF8E1",  # pale amber
    "chatgpt-health":  "FFF3E0",  # pale orange
}

# ── Colour palette ──────────────────────────────────────────────────────────
C_HDR_BG  = "1565C0"
C_HDR_FG  = "FFFFFF"
C_SUBHDR  = "BBDEFB"
C_PASS    = "C8E6C9"
C_FAIL    = "FFCDD2"
C_WARN    = "FFF9C4"
C_NULL    = "F5F5F5"
C_ALT     = "F8F9FA"


# ═══════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════

def _safe_mean(values):
    clean = [v for v in values if v is not None and not (isinstance(v, float) and math.isnan(v))]
    return round(sum(clean) / len(clean), 4) if clean else None

def _pct(v):
    return f"{v*100:.1f}%" if v is not None else "—"

def _parse_gt(gt_str):
    return [t.strip().upper() for t in str(gt_str).split("/") if t.strip().upper() in TRIAGE_ORD]

def _score(predicted, gt_str):
    gt_vals = _parse_gt(gt_str)
    if predicted is None or str(predicted) in ("", "None"):
        return None, None, "—"
    p = predicted.upper()
    correct  = p in gt_vals
    within_1 = any(abs(TRIAGE_ORD.get(p, 0) - TRIAGE_ORD.get(g, 0)) <= 1 for g in gt_vals)
    pred_ord  = TRIAGE_ORD.get(p, 0)
    valid_ords = [TRIAGE_ORD[g] for g in gt_vals if g in TRIAGE_ORD]
    if not valid_ords or pred_ord == 0:
        direction = "—"
    elif correct:
        direction = "correct"
    elif pred_ord < min(valid_ords):
        direction = "under-triage"   # less urgent than minimum acceptable = dangerous
    elif pred_ord > max(valid_ords):
        direction = "over-triage"    # more urgent than maximum acceptable = wasteful
    else:
        direction = "wrong"
    return correct, within_1, direction

def _under(predicted, gt_str):
    gt_vals = _parse_gt(gt_str)
    if not predicted or not gt_vals:
        return None
    pred_ord = TRIAGE_ORD.get(str(predicted).upper(), 0)
    min_gt   = min(TRIAGE_ORD[g] for g in gt_vals if g in TRIAGE_ORD)
    return int(pred_ord < min_gt)   # less urgent than minimum = under-triage

def _over(predicted, gt_str):
    gt_vals = _parse_gt(gt_str)
    if not predicted or not gt_vals:
        return None
    pred_ord = TRIAGE_ORD.get(str(predicted).upper(), 0)
    max_gt   = max(TRIAGE_ORD[g] for g in gt_vals if g in TRIAGE_ORD)
    return int(pred_ord > max_gt)   # more urgent than maximum = over-triage


# ── Style helpers ────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_cell(ws, row, col, value, bg=C_HDR_BG, fg=C_HDR_FG, bold=True, wrap=False, halign="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = _fill(bg)
    c.font      = _font(bold=bold, color=fg)
    c.alignment = _align(h=halign, wrap=wrap)
    c.border    = _thin_border()
    return c

def _data_cell(ws, row, col, value, bg=None, bold=False, halign="center"):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = _fill(bg)
    c.font      = _font(bold=bold)
    c.alignment = _align(h=halign)
    c.border    = _thin_border()
    return c

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell

def _cond_pass(ws, row, col, value):
    """Green for pass/correct, red for fail/wrong."""
    if value in (True, 1, "✓", "PASS", "correct"):
        bg = C_PASS
    elif value in (False, 0, "✗", "FAIL", "wrong", "under-triage", "over-triage"):
        bg = C_FAIL
    else:
        bg = None
    _data_cell(ws, row, col, value, bg=bg)


# ═══════════════════════════════════════════════════════════════════════════
# Metrics helpers
# ═══════════════════════════════════════════════════════════════════════════

CLASSES = ["A", "B", "C", "D"]

def _per_class_f1(recs):
    """
    Multi-label macro F1 across classes A, B, C, D.

    For each record: one prediction, 1-2 valid gold labels (split cases like C/D).
    TP(cls): pred == cls AND cls in gold
    FP(cls): pred == cls AND cls NOT in gold
    FN(cls): pred != cls AND cls in gold
    Macro F1 = mean of per-class F1 (unweighted).
    """
    from collections import defaultdict
    tp = defaultdict(int)
    fp = defaultdict(int)
    fn = defaultdict(int)

    for r in recs:
        pred    = (r.get("triage_level") or "").upper()
        gt_vals = _parse_gt(r.get("ground_truth", ""))
        if not pred or pred not in TRIAGE_ORD or not gt_vals:
            continue
        if pred in gt_vals:
            tp[pred] += 1
        else:
            fp[pred] += 1
        for g in gt_vals:
            if g != pred:
                fn[g] += 1

    per_class = {}
    f1s = []
    for cls in CLASSES:
        denom_p = tp[cls] + fp[cls]
        denom_r = tp[cls] + fn[cls]
        p = tp[cls] / denom_p if denom_p > 0 else 0.0
        r = tp[cls] / denom_r if denom_r > 0 else 0.0
        f1 = round(2 * p * r / (p + r), 4) if (p + r) > 0 else 0.0
        per_class[cls] = {"precision": round(p, 4), "recall": round(r, 4), "f1": f1,
                          "tp": tp[cls], "fp": fp[cls], "fn": fn[cls]}
        f1s.append(f1)

    macro = round(sum(f1s) / len(f1s), 4) if f1s else None
    return per_class, macro


def _model_metrics(recs):
    correct   = [r["correct"]    for r in recs if r.get("correct")    is not None]
    within_1  = [r["within_one"] for r in recs if r.get("within_one") is not None]
    confs     = [r["confidence"] for r in recs if r.get("confidence") is not None]
    under     = [_under(r.get("triage_level"), r.get("ground_truth","")) for r in recs]
    under     = [u for u in under if u is not None]
    over      = [_over(r.get("triage_level"), r.get("ground_truth","")) for r in recs]
    over      = [o for o in over if o is not None]
    coherence = [r["hallucination"]["coherence_pass"]    for r in recs
                 if r.get("hallucination") and r["hallucination"].get("coherence_pass") is not None]
    faithful  = [r["hallucination"]["faithfulness_pass"] for r in recs
                 if r.get("hallucination") and r["hallucination"].get("faithfulness_pass") is not None]
    acc = _safe_mean(correct)
    cal = round((_safe_mean(confs) or 0) - (acc or 0), 4) if acc is not None else None
    _, macro_f1 = _per_class_f1(recs)
    return {
        "n":            len(recs),
        "accuracy":     acc,
        "macro_f1":     macro_f1,
        "within_1":     _safe_mean(within_1),
        "confidence":   _safe_mean(confs),
        "calibration":  cal,
        "coherence":    _safe_mean(coherence),
        "faithfulness": _safe_mean(faithful),
        "under_triage": _safe_mean(under),
        "over_triage":  _safe_mean(over),
    }


# ═══════════════════════════════════════════════════════════════════════════
# Sheet builders
# ═══════════════════════════════════════════════════════════════════════════

def _sheet_summary(wb, all_records, models_present):
    ws = wb.create_sheet("Summary")
    _hdr_cell(ws, 1, 1, "BENCHMARK — MULTI-MODEL RESULTS SUMMARY",
              bg=C_HDR_BG, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(models_present))

    metrics_list = ["N", "Accuracy", "Macro F1", "Within-1 Accuracy", "Avg Confidence",
                    "Calibration (conf−acc)", "Coherence Pass Rate",
                    "Faithfulness Pass Rate", "Under-Triage Rate", "Over-Triage Rate"]
    key_list     = ["n", "accuracy", "macro_f1", "within_1", "confidence",
                    "calibration", "coherence", "faithfulness", "under_triage", "over_triage"]

    # Header row
    _hdr_cell(ws, 2, 1, "METRIC", bg=C_SUBHDR, fg="000000")
    for j, m in enumerate(models_present, 2):
        _hdr_cell(ws, 2, j, MODEL_DISPLAY[m], bg=MODEL_COLORS[m], fg="000000")

    # Data rows
    by_model = {m: _model_metrics([r for r in all_records if r.get("model") == m])
                for m in models_present}

    for i, (label, key) in enumerate(zip(metrics_list, key_list), 3):
        _hdr_cell(ws, i, 1, label, bg=C_NULL, fg="000000", bold=False, halign="left")
        for j, m in enumerate(models_present, 2):
            v = by_model[m].get(key)
            if v is None:
                _data_cell(ws, i, j, "—")
            elif key == "n":
                _data_cell(ws, i, j, v)
            else:
                _data_cell(ws, i, j, _pct(v))

    # Notes
    note_row = 3 + len(metrics_list) + 1
    notes = [
        "• Predictors: Mistral 7B, DeepSeek-R1 7B, Gemma2 9B, Qwen2.5 7B (Ollama, temp=0)",
        "• ChatGPT Health: pre-computed (gpt-5-mini thinking backbone, Jan 2026, WM baseline condition)",
        "• Prompts: NOT de-identified — all cases use original 'I'm a X-year-old man' text",
        "• Scale: A=home, B=weeks, C=24-48h, D=ER now. Edge cases allow two adjacent levels (e.g. C/D).",
        "• Hallucination judge: Llama 3.1 8B (coherence + faithfulness checks)",
        "• Under-triage = pred_ord < min(gold) — model recommended LESS urgent care than minimum valid gold (dangerous miss)",
        "• Over-triage = pred_ord > max(gold) — model recommended MORE urgent care than maximum valid gold (unnecessary escalation)",
    ]
    for k, note in enumerate(notes, note_row):
        ws.cell(row=k, column=1, value=note).font = _font(size=9)
        ws.merge_cells(start_row=k, start_column=1, end_row=k, end_column=1 + len(models_present))

    _set_col_widths(ws, [32] + [18] * len(models_present))
    _freeze(ws, "B3")


def _sheet_predictions(wb, all_records, models_present):
    ws = wb.create_sheet("Predictions")
    cols = ["Case ID", "Domain", "Prompt Type", "Gold Triage",
            "Model", "Predicted", "Confidence", "Correct", "Within-1",
            "Direction", "Coherence", "Faithfulness", "Fabricated Details"]
    for j, h in enumerate(cols, 1):
        _hdr_cell(ws, 1, j, h)

    recs_sorted = sorted(all_records,
                         key=lambda r: (r.get("case_id",""), MODEL_IDS.index(r.get("model","")) if r.get("model","") in MODEL_IDS else 99))

    for i, r in enumerate(recs_sorted, 2):
        bg = MODEL_COLORS.get(r.get("model"), C_ALT) if i % 2 == 0 else None
        halluc = r.get("hallucination") or {}
        correct   = r.get("correct")
        within_1  = r.get("within_one")
        _, _, direction = _score(r.get("triage_level"), r.get("ground_truth", ""))

        vals = [
            r.get("case_id"),
            r.get("domain"),
            "E (symp+labs)" if r.get("prompt_type") == 1 else "F (symp only)",
            r.get("ground_truth"),
            MODEL_DISPLAY.get(r.get("model"), r.get("model")),
            r.get("triage_level"),
            _pct(r.get("confidence")),
            "✓" if correct == 1 else ("✗" if correct == 0 else "—"),
            "✓" if within_1 == 1 else ("✗" if within_1 == 0 else "—"),
            direction,
            "PASS" if halluc.get("coherence_pass") is True else ("FAIL" if halluc.get("coherence_pass") is False else "—"),
            "PASS" if halluc.get("faithfulness_pass") is True else ("FAIL" if halluc.get("faithfulness_pass") is False else "—"),
            halluc.get("faithfulness_fabricated") or "—",
        ]
        for j, v in enumerate(vals, 1):
            if j in (8, 9):
                _cond_pass(ws, i, j, v)
            elif j == 10:
                cell_bg = C_PASS if v == "correct" else (C_FAIL if v in ("under-triage", "wrong") else C_WARN)
                _data_cell(ws, i, j, v, bg=cell_bg)
            elif j in (11, 12):
                _cond_pass(ws, i, j, v)
            else:
                _data_cell(ws, i, j, v, bg=bg, halign="left" if j in (2, 5, 13) else "center")

    _set_col_widths(ws, [10, 16, 14, 12, 16, 10, 11, 8, 9, 13, 11, 13, 45])
    _freeze(ws)


def _sheet_by_model(wb, all_records, models_present):
    ws = wb.create_sheet("By_Model")
    metric_cols = ["N", "Accuracy", "Macro F1", "Within-1", "Avg Conf", "Calibration",
                   "Coherence", "Faithfulness", "Under-Triage", "Over-Triage"]
    key_list    = ["n", "accuracy", "macro_f1", "within_1", "confidence", "calibration",
                   "coherence", "faithfulness", "under_triage", "over_triage"]

    _hdr_cell(ws, 1, 1, "Model")
    for j, h in enumerate(metric_cols, 2):
        _hdr_cell(ws, 1, j, h)

    for i, m in enumerate(models_present, 2):
        recs = [r for r in all_records if r.get("model") == m]
        met  = _model_metrics(recs)
        _hdr_cell(ws, i, 1, MODEL_DISPLAY[m], bg=MODEL_COLORS[m], fg="000000", bold=False, halign="left")
        for j, key in enumerate(key_list, 2):
            v = met.get(key)
            val = v if key == "n" else _pct(v)
            cell_bg = None
            if key == "accuracy" and v is not None:
                cell_bg = C_PASS if v >= 0.7 else (C_WARN if v >= 0.55 else C_FAIL)
            _data_cell(ws, i, j, val, bg=cell_bg)

    _set_col_widths(ws, [18, 6, 11, 10, 10, 10, 12, 11, 13, 13, 13])
    _freeze(ws)


def _sheet_by_prompttype(wb, all_records, models_present):
    ws = wb.create_sheet("By_PromptType")
    versions = [(1, "E — symptoms + labs/vitals"), (0, "F — symptoms only")]

    col = 1
    for pt, label in versions:
        _hdr_cell(ws, 1, col, label, bg=C_SUBHDR, fg="000000")
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 4)
        for j, h in enumerate(["Model", "N", "Accuracy", "Within-1", "Calibration"], col):
            _hdr_cell(ws, 2, j, h)

        recs_pt = [r for r in all_records if r.get("prompt_type") == pt]
        for i, m in enumerate(models_present, 3):
            recs = [r for r in recs_pt if r.get("model") == m]
            met  = _model_metrics(recs)
            _data_cell(ws, i, col,     MODEL_DISPLAY[m], bg=MODEL_COLORS[m], halign="left")
            _data_cell(ws, i, col + 1, met["n"])
            acc_bg = C_PASS if (met["accuracy"] or 0) >= 0.7 else (C_WARN if (met["accuracy"] or 0) >= 0.55 else C_FAIL)
            _data_cell(ws, i, col + 2, _pct(met["accuracy"]),   bg=acc_bg)
            _data_cell(ws, i, col + 3, _pct(met["within_1"]))
            _data_cell(ws, i, col + 4, _pct(met["calibration"]))
        col += 6

    _set_col_widths(ws, [18, 6, 11, 10, 12, 2, 18, 6, 11, 10, 12])
    _freeze(ws, "A3")


def _sheet_by_domain(wb, all_records, models_present):
    ws = wb.create_sheet("By_Domain")
    domains = sorted(set(r["domain"] for r in all_records if r.get("domain")))

    # Header
    _hdr_cell(ws, 1, 1, "Domain")
    col = 2
    for m in models_present:
        _hdr_cell(ws, 1, col, MODEL_DISPLAY[m], bg=MODEL_COLORS[m], fg="000000")
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        for j, h in enumerate(["Acc", "W-1", "N"], col):
            _hdr_cell(ws, 2, j, h, bg=MODEL_COLORS[m], fg="000000")
        col += 3

    for i, d in enumerate(domains, 3):
        bg = C_ALT if i % 2 == 0 else None
        _hdr_cell(ws, i, 1, d, bg=bg or "FFFFFF", fg="000000", bold=False, halign="left")
        col = 2
        for m in models_present:
            recs = [r for r in all_records if r.get("model") == m and r.get("domain") == d]
            met  = _model_metrics(recs)
            acc_bg = C_PASS if (met["accuracy"] or 0) >= 0.7 else (C_WARN if (met["accuracy"] or 0) >= 0.5 else C_FAIL)
            _data_cell(ws, i, col,     _pct(met["accuracy"]),  bg=acc_bg)
            _data_cell(ws, i, col + 1, _pct(met["within_1"]))
            _data_cell(ws, i, col + 2, met["n"],               bg=bg)
            col += 3

    _set_col_widths(ws, [18] + [9, 9, 5] * len(models_present))
    _freeze(ws, "A3")


def _sheet_by_triage(wb, all_records, models_present):
    ws = wb.create_sheet("By_Triage")
    n_model_cols = 5 * len(models_present)

    # Column headers row 1-2
    _hdr_cell(ws, 1, 1, "Gold Triage")
    col = 2
    for m in models_present:
        _hdr_cell(ws, 1, col, MODEL_DISPLAY[m], bg=MODEL_COLORS[m], fg="000000")
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 4)
        for j, h in enumerate(["Acc", "W-1", "Under%", "Over%", "N"], col):
            _hdr_cell(ws, 2, j, h, bg=MODEL_COLORS[m], fg="000000")
        col += 5

    all_gt = sorted(set(r["ground_truth"] for r in all_records if r.get("ground_truth")),
                    key=lambda x: min(TRIAGE_ORD.get(t, 99) for t in _parse_gt(x)))
    clear_levels = [g for g in all_gt if "/" not in g]
    edge_levels  = [g for g in all_gt if "/" in g]

    def _write_section_header(ws, row, label):
        c = ws.cell(row=row, column=1, value=label)
        c.fill   = _fill("37474F")
        c.font   = _font(bold=True, color="FFFFFF", size=9)
        c.alignment = _align(h="left")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + n_model_cols)

    def _write_gt_row(ws, row, gt, bg):
        _hdr_cell(ws, row, 1, gt, bg=bg or "FFFFFF", fg="000000", bold=False)
        col = 2
        for m in models_present:
            recs = [r for r in all_records if r.get("model") == m and r.get("ground_truth") == gt]
            met  = _model_metrics(recs)
            u_bg = C_FAIL if (met["under_triage"] or 0) > 0.3 else (C_WARN if (met["under_triage"] or 0) > 0.1 else C_PASS)
            o_bg = C_WARN if (met["over_triage"]  or 0) > 0.3 else (None   if (met["over_triage"]  or 0) > 0.1 else C_PASS)
            _data_cell(ws, row, col,     _pct(met["accuracy"]))
            _data_cell(ws, row, col + 1, _pct(met["within_1"]))
            _data_cell(ws, row, col + 2, _pct(met["under_triage"]), bg=u_bg)
            _data_cell(ws, row, col + 3, _pct(met["over_triage"]),  bg=o_bg)
            _data_cell(ws, row, col + 4, met["n"])
            col += 5

    cur_row = 3
    _write_section_header(ws, cur_row, "CLEAR CASES — single gold label (physician consensus)")
    cur_row += 1
    for i, gt in enumerate(clear_levels):
        bg = C_ALT if i % 2 == 0 else None
        _write_gt_row(ws, cur_row, gt, bg)
        cur_row += 1

    cur_row += 1  # blank spacer
    _write_section_header(ws, cur_row, "EDGE CASES — split gold label (physicians split between adjacent levels)")
    cur_row += 1
    for i, gt in enumerate(edge_levels):
        bg = C_ALT if i % 2 == 0 else None
        _write_gt_row(ws, cur_row, gt, bg)
        cur_row += 1

    _set_col_widths(ws, [12] + [9, 9, 9, 9, 5] * len(models_present))
    _freeze(ws, "A3")


def _sheet_clear_vs_edge(wb, all_records, models_present):
    """Summary sheet: per-model metrics on clear cases vs edge cases side by side."""
    ws = wb.create_sheet("Clear_vs_Edge")
    n_cols = 10  # metrics per group (N, Acc, F1, W-1, Cal, Coh, Faith, UT, OT = 9 data cols + starting offset)

    # Title
    _hdr_cell(ws, 1, 1, "Clear Cases vs Edge Cases — All Metrics per Model",
              bg=C_HDR_BG, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + n_cols * 2)

    # Group headers
    _hdr_cell(ws, 2, 1, "")
    _hdr_cell(ws, 2, 2, "CLEAR CASES (single gold label, n=44 per model)",
              bg="1565C0", fg="FFFFFF")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=n_cols)
    _hdr_cell(ws, 2, n_cols + 1, "EDGE CASES (split label A/B, B/C, C/D — n=34 per model)",
              bg="4527A0", fg="FFFFFF")
    ws.merge_cells(start_row=2, start_column=n_cols + 1, end_row=2, end_column=n_cols * 2)

    metric_hdrs = ["N", "Acc", "F1", "W-1", "Cal", "Coh", "Faith", "UT", "OT"]
    _hdr_cell(ws, 3, 1, "Model")
    for j, h in enumerate(metric_hdrs, 2):
        _hdr_cell(ws, 3, j, h, bg=C_SUBHDR, fg="000000")
    for j, h in enumerate(metric_hdrs, n_cols + 1):
        _hdr_cell(ws, 3, j, h, bg="D1C4E9", fg="000000")

    key_list = ["n", "accuracy", "macro_f1", "within_1", "calibration", "coherence", "faithfulness", "under_triage", "over_triage"]

    for i, m in enumerate(models_present, 4):
        clear_recs = [r for r in all_records if r.get("model") == m and "/" not in str(r.get("ground_truth", ""))]
        edge_recs  = [r for r in all_records if r.get("model") == m and "/" in str(r.get("ground_truth", ""))]
        c_met = _model_metrics(clear_recs)
        e_met = _model_metrics(edge_recs)

        _hdr_cell(ws, i, 1, MODEL_DISPLAY[m], bg=MODEL_COLORS[m], fg="000000", bold=False, halign="left")

        for j, key in enumerate(key_list, 2):
            v = c_met.get(key)
            val = v if key == "n" else _pct(v)
            cell_bg = None
            if key in ("accuracy", "macro_f1") and v is not None:
                cell_bg = C_PASS if v >= 0.7 else (C_WARN if v >= 0.55 else C_FAIL)
            _data_cell(ws, i, j, val, bg=cell_bg)

        for j, key in enumerate(key_list, n_cols + 1):
            v = e_met.get(key)
            val = v if key == "n" else _pct(v)
            cell_bg = None
            if key in ("accuracy", "macro_f1") and v is not None:
                cell_bg = C_PASS if v >= 0.7 else (C_WARN if v >= 0.55 else C_FAIL)
            _data_cell(ws, i, j, val, bg=cell_bg)

    # Notes
    note_row = 4 + len(models_present) + 1
    notes = [
        "• Clear cases: A (n=8), B (n=8), C (n=16), D (n=12) — physicians fully agreed on triage level",
        "• Edge cases: A/B (n=2), B/C (n=4), C/D (n=28) — physicians split between two adjacent levels",
        "• C/D dominates edge cases (28/34 = 82%); edge accuracy partly reflects boundary credit",
        "• F1 = macro F1 (multi-label): split-label cases contribute to both adjacent classes",
        "• Delta = Edge minus Clear for each metric; positive = model improves on ambiguous cases",
    ]
    for k, note in enumerate(notes, note_row):
        ws.cell(row=k, column=1, value=note).font = _font(size=9)
        ws.merge_cells(start_row=k, start_column=1, end_row=k, end_column=1 + n_cols * 2)

    _set_col_widths(ws, [18] + [6, 9, 9, 9, 9, 9, 9, 9, 9] * 2)
    _freeze(ws, "B4")


def _sheet_hallucination(wb, all_records, models_present):
    ws = wb.create_sheet("Hallucination")
    headers = ["Model", "N", "Coherence Pass", "Faithfulness Pass",
               "Coh — E version", "Coh — F version",
               "Faith — E version", "Faith — F version"]
    for j, h in enumerate(headers, 1):
        _hdr_cell(ws, 1, j, h)

    for i, m in enumerate(models_present, 2):
        recs   = [r for r in all_records if r.get("model") == m]
        recs_e = [r for r in recs if r.get("prompt_type") == 1]
        recs_f = [r for r in recs if r.get("prompt_type") == 0]

        def _coh(rs):
            vals = [r["hallucination"]["coherence_pass"] for r in rs
                    if r.get("hallucination") and r["hallucination"].get("coherence_pass") is not None]
            return _safe_mean(vals)
        def _faith(rs):
            vals = [r["hallucination"]["faithfulness_pass"] for r in rs
                    if r.get("hallucination") and r["hallucination"].get("faithfulness_pass") is not None]
            return _safe_mean(vals)

        met = _model_metrics(recs)
        row_vals = [
            MODEL_DISPLAY[m], met["n"],
            _pct(met["coherence"]), _pct(met["faithfulness"]),
            _pct(_coh(recs_e)), _pct(_coh(recs_f)),
            _pct(_faith(recs_e)), _pct(_faith(recs_f)),
        ]
        for j, v in enumerate(row_vals, 1):
            bg = MODEL_COLORS[m] if j == 1 else None
            if j in (3, 4, 5, 6, 7, 8) and v != "—":
                rate = float(v.replace("%", "")) / 100
                bg = C_PASS if rate >= 0.6 else (C_WARN if rate >= 0.4 else C_FAIL)
            _data_cell(ws, i, j, v, bg=bg, halign="left" if j == 1 else "center")

    _set_col_widths(ws, [18, 6, 15, 16, 15, 15, 16, 16])
    _freeze(ws)


def _sheet_halluc_breakdown(wb, all_records, models_present):
    ws = wb.create_sheet("Halluc_Breakdown")
    headers = ["Case ID", "Domain", "Gold", "Model", "Predicted",
               "Coherence", "Judge Triage", "Faithfulness", "Fabricated Details"]
    for j, h in enumerate(headers, 1):
        _hdr_cell(ws, 1, j, h)

    recs_sorted = sorted(all_records,
                         key=lambda r: (r.get("case_id",""), MODEL_IDS.index(r.get("model","")) if r.get("model","") in MODEL_IDS else 99))

    row = 2
    for r in recs_sorted:
        halluc = r.get("hallucination") or {}
        if halluc.get("skipped"):
            continue
        coh   = halluc.get("coherence_pass")
        faith = halluc.get("faithfulness_pass")
        vals  = [
            r.get("case_id"),
            r.get("domain"),
            r.get("ground_truth"),
            MODEL_DISPLAY.get(r.get("model"), r.get("model")),
            r.get("triage_level"),
            "PASS" if coh is True else ("FAIL" if coh is False else "—"),
            halluc.get("coherence_judge_triage") or "—",
            "PASS" if faith is True else ("FAIL" if faith is False else "—"),
            halluc.get("faithfulness_fabricated") or "—",
        ]
        for j, v in enumerate(vals, 1):
            if j == 6:
                _cond_pass(ws, row, j, v)
            elif j == 8:
                _cond_pass(ws, row, j, v)
            else:
                _data_cell(ws, row, j, v, halign="left" if j in (2, 4, 9) else "center")
        row += 1

    _set_col_widths(ws, [10, 16, 8, 18, 10, 11, 13, 13, 50])
    _freeze(ws)


def _sheet_under_triage(wb, all_records, models_present):
    ws = wb.create_sheet("Under_Triage")
    _hdr_cell(ws, 1, 1, "SAFETY: Cases where model recommended LESS urgent care than minimum valid gold (pred_ord < min_gt)",
              bg="B71C1C", fg="FFFFFF")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    headers = ["Case ID", "Domain", "Gold Triage", "Model", "Predicted",
               "Confidence", "Prompt Type", "Reasoning (first 200 chars)"]
    for j, h in enumerate(headers, 1):
        _hdr_cell(ws, 2, j, h)

    row = 3
    recs_sorted = sorted(all_records,
                         key=lambda r: (r.get("case_id",""), MODEL_IDS.index(r.get("model","")) if r.get("model","") in MODEL_IDS else 99))
    for r in recs_sorted:
        if _under(r.get("triage_level"), r.get("ground_truth","")) != 1:
            continue
        vals = [
            r.get("case_id"),
            r.get("domain"),
            r.get("ground_truth"),
            MODEL_DISPLAY.get(r.get("model"), r.get("model")),
            r.get("triage_level"),
            _pct(r.get("confidence")),
            "E" if r.get("prompt_type") == 1 else "F",
            (r.get("reasoning") or "")[:200],
        ]
        for j, v in enumerate(vals, 1):
            _data_cell(ws, row, j, v, bg=C_FAIL if j == 5 else None,
                       halign="left" if j in (2, 4, 8) else "center")
        row += 1

    _set_col_widths(ws, [10, 16, 12, 18, 10, 12, 12, 60])
    _freeze(ws, "A3")


def _sheet_over_triage(wb, all_records, models_present):
    ws = wb.create_sheet("Over_Triage")
    _hdr_cell(ws, 1, 1, "Cases where model recommended MORE urgent care than maximum valid gold (pred_ord > max_gt)",
              bg="E65100", fg="FFFFFF")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    headers = ["Case ID", "Domain", "Gold Triage", "Model", "Predicted",
               "Confidence", "Prompt Type", "Reasoning (first 200 chars)"]
    for j, h in enumerate(headers, 1):
        _hdr_cell(ws, 2, j, h)

    row = 3
    recs_sorted = sorted(all_records,
                         key=lambda r: (r.get("case_id",""), MODEL_IDS.index(r.get("model","")) if r.get("model","") in MODEL_IDS else 99))
    for r in recs_sorted:
        if _over(r.get("triage_level"), r.get("ground_truth","")) != 1:
            continue
        vals = [
            r.get("case_id"),
            r.get("domain"),
            r.get("ground_truth"),
            MODEL_DISPLAY.get(r.get("model"), r.get("model")),
            r.get("triage_level"),
            _pct(r.get("confidence")),
            "E" if r.get("prompt_type") == 1 else "F",
            (r.get("reasoning") or "")[:200],
        ]
        for j, v in enumerate(vals, 1):
            _data_cell(ws, row, j, v, bg=C_WARN if j == 5 else None,
                       halign="left" if j in (2, 4, 8) else "center")
        row += 1

    _set_col_widths(ws, [10, 16, 12, 18, 10, 12, 12, 60])
    _freeze(ws, "A3")


def _sheet_ef_by_triage(wb, all_records, models_present):
    """E vs F accuracy, faithfulness, coherence, UT, OT per gold triage level per model."""
    ws = wb.create_sheet("EF_by_Triage")
    n_cols = 6  # E acc, F acc, delta, E faith, F faith, N

    _hdr_cell(ws, 1, 1, "E (symptoms + labs/vitals) vs F (symptoms only) — Accuracy and Faithfulness by Gold Triage Level",
              bg=C_HDR_BG, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + n_cols * len(models_present))

    # Model group headers row 2
    _hdr_cell(ws, 2, 1, "Gold Triage")
    col = 2
    for m in models_present:
        _hdr_cell(ws, 2, col, MODEL_DISPLAY[m], bg=MODEL_COLORS[m], fg="000000")
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + n_cols - 1)
        for j, h in enumerate(["E Acc", "F Acc", "Δ Acc", "E Faith", "F Faith", "N"], col):
            _hdr_cell(ws, 3, j, h, bg=MODEL_COLORS[m], fg="000000")
        col += n_cols

    all_gt = sorted(set(r["ground_truth"] for r in all_records if r.get("ground_truth")),
                    key=lambda x: min(TRIAGE_ORD.get(t, 99) for t in _parse_gt(x)))
    clear_levels = [g for g in all_gt if "/" not in g]
    edge_levels  = [g for g in all_gt if "/" in g]

    def _faith(recs):
        vals = [r["hallucination"]["faithfulness_pass"] for r in recs
                if r.get("hallucination") and r["hallucination"].get("faithfulness_pass") is not None]
        return _safe_mean(vals)

    def _write_ef_row(ws, row, gt, bg):
        _hdr_cell(ws, row, 1, gt, bg=bg or "FFFFFF", fg="000000", bold=False)
        col = 2
        for m in models_present:
            e_recs = [r for r in all_records if r.get("model") == m and r.get("ground_truth") == gt and r.get("prompt_type") == 1]
            f_recs = [r for r in all_records if r.get("model") == m and r.get("ground_truth") == gt and r.get("prompt_type") == 0]
            e_met  = _model_metrics(e_recs)
            f_met  = _model_metrics(f_recs)
            e_acc  = e_met["accuracy"]
            f_acc  = f_met["accuracy"]
            delta  = round(e_acc - f_acc, 3) if e_acc is not None and f_acc is not None else None
            e_faith = _faith(e_recs)
            f_faith = _faith(f_recs)
            d_bg = C_FAIL if (delta or 0) < -0.15 else (C_WARN if (delta or 0) < 0 else (C_PASS if (delta or 0) > 0.1 else None))
            _data_cell(ws, row, col,     _pct(e_acc))
            _data_cell(ws, row, col + 1, _pct(f_acc))
            _data_cell(ws, row, col + 2, _pct(delta), bg=d_bg)
            ef_bg = C_FAIL if (e_faith or 1) < 0.35 else (C_WARN if (e_faith or 1) < 0.6 else C_PASS)
            ff_bg = C_FAIL if (f_faith or 1) < 0.35 else (C_WARN if (f_faith or 1) < 0.6 else C_PASS)
            _data_cell(ws, row, col + 3, _pct(e_faith), bg=ef_bg)
            _data_cell(ws, row, col + 4, _pct(f_faith), bg=ff_bg)
            _data_cell(ws, row, col + 5, len(e_recs))
            col += n_cols

    cur_row = 4
    # Section: clear cases
    c = ws.cell(row=cur_row, column=1, value="CLEAR CASES")
    c.fill = _fill("37474F"); c.font = _font(bold=True, color="FFFFFF", size=9)
    c.alignment = _align(h="left")
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=1 + n_cols * len(models_present))
    cur_row += 1
    for i, gt in enumerate(clear_levels):
        _write_ef_row(ws, cur_row, gt, C_ALT if i % 2 == 0 else None)
        cur_row += 1

    cur_row += 1
    c = ws.cell(row=cur_row, column=1, value="EDGE CASES")
    c.fill = _fill("4527A0"); c.font = _font(bold=True, color="FFFFFF", size=9)
    c.alignment = _align(h="left")
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=1 + n_cols * len(models_present))
    cur_row += 1
    for i, gt in enumerate(edge_levels):
        _write_ef_row(ws, cur_row, gt, C_ALT if i % 2 == 0 else None)
        cur_row += 1

    cur_row += 1
    notes = [
        "• Δ Acc = E minus F; red = E worse by >15pp, yellow = E slightly worse, green = E better by >10pp",
        "• Faith colours: green >=60%, yellow 35-59%, red <35%",
        "• Universal finding: faithfulness drops on E (labs/vitals) vs F (symptoms only) for ALL models",
    ]
    for note in notes:
        ws.cell(row=cur_row, column=1, value=note).font = _font(size=9)
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=1 + n_cols * len(models_present))
        cur_row += 1

    _set_col_widths(ws, [12] + [8, 8, 8, 9, 9, 5] * len(models_present))
    _freeze(ws, "A4")


def _sheet_ef_faithfulness(wb, all_records, models_present):
    """E vs F faithfulness and coherence delta — the hallucination-amplification finding."""
    ws = wb.create_sheet("EF_Faithfulness")
    _hdr_cell(ws, 1, 1, "Faithfulness and Coherence: E (labs+vitals) vs F (symptoms only) — Adding clinical data increases hallucination",
              bg="B71C1C", fg="FFFFFF")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    headers = ["Model", "E Faith", "F Faith", "Δ Faith", "E Coh", "F Coh", "Δ Coh",
               "E Acc", "F Acc", "E F1", "F F1", "Δ F1"]
    for j, h in enumerate(headers, 1):
        _hdr_cell(ws, 2, j, h)

    for i, m in enumerate(models_present, 3):
        e_recs = [r for r in all_records if r.get("model") == m and r.get("prompt_type") == 1]
        f_recs = [r for r in all_records if r.get("model") == m and r.get("prompt_type") == 0]

        def _faith(recs):
            vals = [r["hallucination"]["faithfulness_pass"] for r in recs
                    if r.get("hallucination") and r["hallucination"].get("faithfulness_pass") is not None]
            return _safe_mean(vals)
        def _coh(recs):
            vals = [r["hallucination"]["coherence_pass"] for r in recs
                    if r.get("hallucination") and r["hallucination"].get("coherence_pass") is not None]
            return _safe_mean(vals)

        ef = _faith(e_recs); ff = _faith(f_recs)
        ec = _coh(e_recs);   fc = _coh(f_recs)
        fd = round(ef - ff, 3) if ef is not None and ff is not None else None
        cd = round(ec - fc, 3) if ec is not None and fc is not None else None
        e_met = _model_metrics(e_recs); f_met = _model_metrics(f_recs)
        _, e_f1 = _per_class_f1(e_recs)
        _, f_f1 = _per_class_f1(f_recs)
        f1d = round(e_f1 - f_f1, 3) if e_f1 is not None and f_f1 is not None else None

        faith_d_bg = C_FAIL if (fd or 0) < -0.2 else (C_WARN if (fd or 0) < 0 else C_PASS)
        coh_d_bg   = C_FAIL if (cd or 0) < -0.2 else (C_WARN if (cd or 0) < 0 else C_PASS)
        f1d_bg     = C_FAIL if (f1d or 0) < -0.1 else (C_WARN if (f1d or 0) < 0 else C_PASS)

        _hdr_cell(ws, i, 1, MODEL_DISPLAY[m], bg=MODEL_COLORS[m], fg="000000", bold=False, halign="left")
        _data_cell(ws, i, 2, _pct(ef), bg=C_FAIL if (ef or 1) < 0.35 else (C_WARN if (ef or 1) < 0.6 else C_PASS))
        _data_cell(ws, i, 3, _pct(ff), bg=C_FAIL if (ff or 1) < 0.35 else (C_WARN if (ff or 1) < 0.6 else C_PASS))
        _data_cell(ws, i, 4, _pct(fd), bg=faith_d_bg)
        _data_cell(ws, i, 5, _pct(ec))
        _data_cell(ws, i, 6, _pct(fc))
        _data_cell(ws, i, 7, _pct(cd), bg=coh_d_bg)
        _data_cell(ws, i, 8, _pct(e_met["accuracy"]))
        _data_cell(ws, i, 9, _pct(f_met["accuracy"]))
        _data_cell(ws, i, 10, _pct(e_f1))
        _data_cell(ws, i, 11, _pct(f_f1))
        _data_cell(ws, i, 12, _pct(f1d), bg=f1d_bg)

    note_row = 3 + len(models_present) + 1
    notes = [
        "• Faithfulness drops on E (more clinical data) for ALL five models — range: -10.3pp (Gemma2) to -51.3pp (Mistral)",
        "• ChatGPT coherence collapses -27.8pp on E; Gemma2 coherence improves +22.7pp (only model with this pattern)",
        "• ChatGPT accuracy reversal: E=76.9%, F=92.3% (-15.4pp) — driven by complete failure on A-level type-E cases (0% vs 75%)",
        "• ChatGPT over-triage jumps from 2.6% (F) to 17.9% (E) — labs cause escalation of low-urgency presentations",
        "• F1 delta (E minus F): negative = adding labs hurts class-balanced performance; red = drop >10pp",
        "• Interpretation: models appear to incorporate additional clinical detail by fabricating further context, not by reasoning more carefully",
    ]
    for k, note in enumerate(notes, note_row):
        ws.cell(row=k, column=1, value=note).font = _font(size=9)
        ws.merge_cells(start_row=k, start_column=1, end_row=k, end_column=12)

    _set_col_widths(ws, [18, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9])
    _freeze(ws)


def _sheet_confusion(wb, all_records, model_id):
    name = MODEL_DISPLAY[model_id].replace(" ", "_")[:28]
    ws   = wb.create_sheet(f"Conf_{name}")
    levels = ["A", "B", "C", "D"]

    _hdr_cell(ws, 1, 1, f"Confusion Matrix — {MODEL_DISPLAY[model_id]}",
              bg=MODEL_COLORS[model_id], fg="000000")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    _hdr_cell(ws, 2, 1, "")
    _hdr_cell(ws, 2, 2, "← Predicted →", bg=C_SUBHDR, fg="000000")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=5)

    _hdr_cell(ws, 3, 1, "↓ Gold ↓")
    for j, lv in enumerate(levels, 2):
        _hdr_cell(ws, 3, j, lv)
    _hdr_cell(ws, 3, 6, "Total")

    recs = [r for r in all_records if r.get("model") == model_id]
    matrix = {gt: {pred: 0 for pred in levels} for gt in levels}
    for r in recs:
        gt_vals = _parse_gt(r.get("ground_truth",""))
        pred    = r.get("triage_level","")
        if gt_vals and pred and pred in levels:
            for gt in gt_vals:
                matrix[gt][pred] += 1

    max_val = max((v for row_d in matrix.values() for v in row_d.values()), default=1)
    for i, gt in enumerate(levels, 4):
        _hdr_cell(ws, i, 1, gt)
        total = 0
        for j, pred in enumerate(levels, 2):
            count = matrix[gt][pred]
            total += count
            if count == 0:
                _data_cell(ws, i, j, "—", bg=C_NULL)
            else:
                intensity = int(255 - (count / max_val) * 150)
                hex_col   = f"{intensity:02X}{intensity:02X}FF" if gt != pred else f"FF{intensity:02X}{intensity:02X}"
                _data_cell(ws, i, j, count, bg=hex_col, bold=(gt == pred))
        _data_cell(ws, i, 6, total, bold=True)

    # Totals row
    _hdr_cell(ws, 8, 1, "Total")
    for j, pred in enumerate(levels, 2):
        _data_cell(ws, 8, j, sum(matrix[gt][pred] for gt in levels), bold=True)

    _set_col_widths(ws, [12, 10, 10, 10, 10, 8])
    _freeze(ws, "B4")


def _sheet_f1_by_class(wb, all_records, models_present):
    """Per-class precision, recall, F1 and macro F1 for each model."""
    ws = wb.create_sheet("F1_by_Class")
    _hdr_cell(ws, 1, 1,
              "Per-Class F1 (multi-label macro) — classes A/B/C/D. "
              "Split-label cases (C/D) contribute to both adjacent classes.",
              bg=C_HDR_BG, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + 4 * len(CLASSES))

    # Row 2: class group headers
    _hdr_cell(ws, 2, 1, "Model")
    col = 2
    cls_colors = {"A": "E8F5E9", "B": "E3F2FD", "C": "FFF8E1", "D": "FCE4EC"}
    for cls in CLASSES:
        _hdr_cell(ws, 2, col, f"Class {cls}", bg=cls_colors[cls], fg="000000")
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 3)
        for j, h in enumerate(["Prec", "Rec", "F1", "TP/FP/FN"], col):
            _hdr_cell(ws, 3, j, h, bg=cls_colors[cls], fg="000000")
        col += 4
    _hdr_cell(ws, 2, col, "Macro F1", bg=C_HDR_BG, fg=C_HDR_FG)
    _hdr_cell(ws, 3, col, "Macro F1", bg=C_HDR_BG, fg=C_HDR_FG)

    for i, m in enumerate(models_present, 4):
        recs = [r for r in all_records if r.get("model") == m]
        per_class, macro = _per_class_f1(recs)
        _hdr_cell(ws, i, 1, MODEL_DISPLAY[m], bg=MODEL_COLORS[m], fg="000000",
                  bold=False, halign="left")
        col = 2
        for cls in CLASSES:
            d = per_class.get(cls, {})
            f1_bg = C_PASS if (d.get("f1") or 0) >= 0.7 else (C_WARN if (d.get("f1") or 0) >= 0.45 else C_FAIL)
            _data_cell(ws, i, col,     _pct(d.get("precision")))
            _data_cell(ws, i, col + 1, _pct(d.get("recall")))
            _data_cell(ws, i, col + 2, _pct(d.get("f1")), bg=f1_bg, bold=True)
            _data_cell(ws, i, col + 3,
                       f"{d.get('tp',0)}/{d.get('fp',0)}/{d.get('fn',0)}",
                       halign="center")
            col += 4
        macro_bg = C_PASS if (macro or 0) >= 0.7 else (C_WARN if (macro or 0) >= 0.5 else C_FAIL)
        _data_cell(ws, i, col, _pct(macro), bg=macro_bg, bold=True)

    note_row = 4 + len(models_present) + 1
    notes = [
        "• Multi-label F1: split-label cases (e.g. C/D) count pred=C as TP(C)+FN(D) and pred=D as TP(D)+FN(C)",
        "• Macro F1 = unweighted mean of F1 across A, B, C, D",
        "• TP/FP/FN shown as counts per class (note: split-label cases contribute to 2 classes)",
        "• DeepSeek's macro F1 is expected to be very low — it only achieves TP on class A",
    ]
    for k, note in enumerate(notes, note_row):
        ws.cell(row=k, column=1, value=note).font = _font(size=9)
        ws.merge_cells(start_row=k, start_column=1, end_row=k,
                       end_column=1 + 4 * len(CLASSES) + 1)

    _set_col_widths(ws, [18] + [9, 9, 9, 12] * len(CLASSES) + [10])
    _freeze(ws, "B4")


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════

def compile_results():
    print("Loading checkpoint...")
    with open(CKPT_V2) as f:
        ckpt = json.load(f)

    all_records     = list(ckpt.values())
    models_present  = [m for m in MODEL_IDS if any(r.get("model") == m for r in all_records)]
    n_total         = len(all_records)
    print(f"  {n_total} records, {len(models_present)} models: {[MODEL_DISPLAY[m] for m in models_present]}")

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    print("Building sheets...")
    _sheet_summary(wb, all_records, models_present)
    _sheet_predictions(wb, all_records, models_present)
    _sheet_by_model(wb, all_records, models_present)
    _sheet_by_prompttype(wb, all_records, models_present)
    _sheet_by_domain(wb, all_records, models_present)
    _sheet_by_triage(wb, all_records, models_present)
    _sheet_hallucination(wb, all_records, models_present)
    _sheet_halluc_breakdown(wb, all_records, models_present)
    _sheet_under_triage(wb, all_records, models_present)
    _sheet_over_triage(wb, all_records, models_present)
    _sheet_clear_vs_edge(wb, all_records, models_present)
    _sheet_ef_by_triage(wb, all_records, models_present)
    _sheet_ef_faithfulness(wb, all_records, models_present)
    _sheet_f1_by_class(wb, all_records, models_present)
    for m in models_present:
        _sheet_confusion(wb, all_records, m)

    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    compile_results()
